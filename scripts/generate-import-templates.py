#!/usr/bin/env python3
"""Generate Excel import templates for rack / server / device / cable imports.

Requires: python3 -m venv .venv-templates && .venv-templates/bin/pip install openpyxl
Run:      .venv-templates/bin/python scripts/generate-import-templates.py
"""

from __future__ import annotations

import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    print("Missing dependency: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
OUT_DIRS = [
    ROOT / "scripts/package/import-templates",
    ROOT / "docs",
]

HEADER_FONT = Font(bold=True)


def save_workbook(wb: Workbook, filename: str) -> None:
    for out_dir in OUT_DIRS:
        out_dir.mkdir(parents=True, exist_ok=True)
        path = out_dir / filename
        wb.save(path)
        print(f"Wrote {path}")


def build_rack_template() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "机柜导入"
    headers = ["机柜编号", "所在机房", "高度(U)", "品牌", "额定功率", "备注", "X坐标", "Y坐标", "Z坐标"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    ws.append(["A01", "示例机房", 42, "华为", 5.5, "核心交换机柜", 0, 0, 0])
    ws.append(["A02", "示例机房", 42, "", "", "应用服务器柜", 600, 0, 0])
    ws.append(["B01", "示例机房", 47, "戴尔", 6.0, "", 0, 600, 0])
    return wb


def build_server_template() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "服务器导入"
    headers = [
        "服务器名称", "管理IP", "资产编号", "设备类型", "设备高度(U)",
        "运行状态", "系统", "负责人", "备注", "所在机柜", "起始U位",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    ws.append([
        "app-web-01", "10.0.10.21", "AST-001", "服务器", 1,
        "正常", "Web业务", "张工", "示例 Web 服务器", "A01", 42,
    ])
    ws.append([
        "app-db-01", "10.0.10.22", "AST-002", "服务器", 2,
        "正常", "数据库", "李工", "2U 数据库服务器", "A02", 41,
    ])
    ws.append([
        "net-core-sw-01", "10.0.10.1", "AST-003", "交换机", 1,
        "正常", "核心网络", "王工", "核心交换机", "A01", 40,
    ])
    return wb


def build_device_template() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "设备落位"
    ws.append(["U位", "[A01]", "A01-设备", "[A02]", "A02-设备"])
    for cell in ws[1]:
        cell.font = HEADER_FONT
    # U42: 1U 交换机；U41-U40: 2U 数据库；U39: 空闲
    ws.append([42, 42, "核心交换机", 42, "存储设备"])
    ws.append([41, 41, "", 41, ""])
    ws.append([40, 40, "", 40, ""])
    ws.append([39, 39, "", 39, "防火墙"])
    ws.merge_cells("C3:C4")
    ws["C3"].value = "2U 数据库"
    return wb


def build_cable_template() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "线缆导入"
    headers = [
        "源设备", "源端口", "源端口类型", "源端口速率",
        "目标设备", "目标端口", "目标端口类型", "目标端口速率",
        "线缆类型", "颜色", "长度", "线路用途",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    ws.append([
        "app-web-01", "GE0/0/1", "RJ45", "1G",
        "net-core-sw-01", "GE0/0/1", "RJ45", "1G",
        "铜缆", "蓝色", "3m", "正常",
    ])
    ws.append([
        "app-db-01", "eth0", "RJ45", "10G",
        "net-core-sw-01", "GE0/0/2", "RJ45", "10G",
        "DAC", "灰色", "1m", "业务网络",
    ])
    return wb


def main() -> None:
    templates = [
        ("机柜导入模板.xlsx", build_rack_template),
        ("服务器导入模板.xlsx", build_server_template),
        ("设备导入模板.xlsx", build_device_template),
        ("线缆导入模板.xlsx", build_cable_template),
    ]
    for filename, builder in templates:
        save_workbook(builder(), filename)


if __name__ == "__main__":
    main()
